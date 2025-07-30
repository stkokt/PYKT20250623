import openpyxl as xl

wb = xl.load_workbook('inlandsprodukt-lange-reihen-xlsx-2180150.xlsx')
ws = wb['2.5'] # Erwerbstätige in Berufsgruppen


# A15:K76

# Einlesen in Liste/ Variante 1
values1 = []
for row in ws.iter_rows(min_row=15, max_row=76, min_col=1, max_col=11):
    tmp = []
    for cell in row:
        tmp.append(cell.value)
    values1.append(tmp.copy())

# print(values1)

# Einlesen in Liste/ Variante 2
values2 = [[cell.value for cell in row] for row in  ws.iter_rows(min_row=15,max_row=76,min_col=1,max_col=11)]

#print(values2)

# Einlesen in Liste/ Variante 3
rangeXL = ws['A15':'K76']

# print(range)

values3 = [[x.value for x in cell] for cell in rangeXL]

# Durch Formatierungen im Exceldokument haben wir Zeilen,
# die keine Werte enthalten:
# print(values3)

# Diese wollen wir entfernen. Der Einfachheit halber schreiben
# wir in eine neue Variable.
values4 = [[x for x in line] for line in values3 if line[0]!=None]

#print(values4)

# 2019 fand eine Revision ab dem Jahr 1991 statt. Für das Jahr 1991 finden sich 
# zwei Datensätze im Datenblatt. Diese beiden Datensätze schreiben wir uns mal in 
# eine weitere Variable, um sie weiter zu verarbeiten.

# Die Jahreszahl findet sich immer am Index 0 eines jeden Datensatzes.
# Wir finden den Index des ersten Datensatzes für 1991 hiermit:
firstIdx_1991 = [values4[n][0]==1991 for n,line in enumerate(values4)].index(True)

# Mit einem List- Slicing isolieren wir die beiden Datensätze:
values1991 = values4[firstIdx_1991:firstIdx_1991+2]

# Vielleicht ist es datentechnisch nicht das richtige Vorgehen, 
# aber wir interpolieren einfach alle Daten (nicht das Jahr) linear:
values1991avg = [int((values1991[0][idx]+values1991[1][idx])/2) if idx != 0 else 1991  for idx, n in enumerate(values1991[0])]
#print(values1991avg)

# Jetzt entfernen wir die beiden Datensätze aus unserer Datenreihe:
del values4[firstIdx_1991:firstIdx_1991+2]
#print(values4)

# ...und setzen unsere interpolierte Datenreihe an diese Stelle:
values4.insert(firstIdx_1991, values1991avg)
print(values4)

# Wir legen in unserem Verzeichnis eine neue Excel- Datei an und speichern sie.
myWB = xl.Workbook('myWorkbook.xlsx')
myWB.save('myWorkbook.xlsx')

# In diese Mappe wollen wir unsere Datenreihen schreiben.
# Wir wählen die Datei ganz professionell über einen Dialog aus:
import tkinter as tk
from tkinter import filedialog

path = filedialog.askopenfilename(
    title="Excel- Dateien suchen und öffnen", # Fenstertitel des Dialogs
    filetypes=(("Excel- Dateien", "*.xlsx"), ("Alle Dateien", "*.*")), # Auswahlmölichkeiten
    # Das Verzeichnis, das im Dialog vorausgewählt wird (Ist natürlich bei jedem anders.)
    initialdir="C:/Users/StefanKoschnik/OneDrive - karriere tutor GmbH/Dokumente/Python/PYKT20250623/PYKT20250623/CSV/Wirtschaft"
)

# Wir laden uns die Datei in den Arbeitsspeicher:
wb1 = xl.load_workbook(path)
# Bei einer neu erstellten Excel- Datei heißt die erste Tabelle 'Sheet'.
ws1 = wb1['Sheet']
# Diese Bezeichnung wollen wir ändern:
ws1.title = "Beschäftigte 1970-2022"

# In diese Tabelle tragen wir unsere Werte ein.
for row_idx, val in enumerate(values4, start=1):
    for col_idx,v in enumerate(val, start=1):
        ws1.cell(row=row_idx, column = col_idx, value=v)

# Für weitere Datenreihen können wir uns eine neue Tabelle anlegen
wb1.create_sheet(title="Zweite Tabelle",index=1) # index=1 : hinter der ersten Tabelle
wb1.save('myWorkbook.xlsx')